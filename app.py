from flask import Flask, render_template, request, send_file, jsonify
from datetime import date, timedelta, datetime
import os
import json
import time, random, tempfile, re, time, io, zipfile, csv
import requests
from dotenv import load_dotenv

# Load local .env before reading API credentials.
load_dotenv()

app = Flask(__name__)

AUTH_URL = os.environ.get("AUTH_URL", "https://devapigee.itnext-dev.com/api/v1/auth/token")
CREATE_QUOTE_URL = os.environ.get("CREATE_QUOTE_URL", "https://devapigee.itnext-dev.com/healthproduct-uat/v1/createQuote")
CREATE_PAYMENT_URL = os.environ.get("CREATE_PAYMENT_URL", "https://devapigee.itnext-dev.com/billing/common-uat/v2/createPayment")
ISSUE_POLICY_URL = os.environ.get("ISSUE_POLICY_URL", "https://devapigee.itnext-dev.com/healthproduct-uat/v1/issuePolicy")
RENEWAL_BASE_QUOTE_URL = os.environ.get("RENEWAL_BASE_QUOTE_URL", "https://devapigee.itnext-dev.com/policy/common-uat/v1/GetRenewalBaseQuoteDetails")
API_CLIENT_ID = os.environ.get("API_CLIENT_ID", "")
API_CLIENT_SECRET = os.environ.get("API_CLIENT_SECRET", "")
PAYMENT_COOKIE = os.environ.get("PAYMENT_COOKIE", "")

ISSUE_POLICY_INITIAL_DELAY = float(
    os.environ.get("ISSUE_POLICY_INITIAL_DELAY", "4")
)
ISSUE_POLICY_MAX_ATTEMPTS = int(
    os.environ.get("ISSUE_POLICY_MAX_ATTEMPTS", "5")
)
ISSUE_POLICY_RETRY_DELAY = float(
    os.environ.get("ISSUE_POLICY_RETRY_DELAY", "3")
)
API_TIMEOUT_SECONDS = int(os.environ.get("API_TIMEOUT_SECONDS", "60"))

_TOKEN_CACHE = {"access_token": None, "expires_at": 0}

RELATIONS = {
    'Self': dict(gender='M', salutation='MR', age=(25,55), height=(160,185), weight=(55,95), proposer='1'),
    'Wife': dict(gender='F', salutation='MRS', age=(23,52), height=(148,175), weight=(45,80), proposer='0'),
    'Husband': dict(gender='M', salutation='MR', age=(25,58), height=(160,188), weight=(58,100), proposer='0'),
    'Father': dict(gender='M', salutation='MR', age=(50,75), height=(155,180), weight=(55,90), proposer='0'),
    'Mother': dict(gender='F', salutation='MRS', age=(48,72), height=(145,170), weight=(45,82), proposer='0'),
    'Son': dict(gender='M', salutation='MASTER', age=(3,24), height=(95,178), weight=(15,75), proposer='0'),
    'Daughter': dict(gender='F', salutation='MISS', age=(3,24), height=(95,170), weight=(15,68), proposer='0'),
}

COVERAGES = {
    "MyHealthCriticalIllness": {
        "Indicator": "0",
        "SumInsured": 0,
        "PlanName": "7",
    },
    "IndividualPersonalAccidentRider": {
        "Indicator": "0",
    },
    "OptimaWellbeing": {
        "Indicator": "0",
    },
    "MyHealthHospitalCashBenefit": {
        "Indicator": "0",
        "IsGlobal": "0",
        "SumInsured": 1000,
    },
    "UnlimitedRestoreRider": {
        "Indicator": "0",
    },
    "ProtectorRider": {
        "Indicator": "0",
    },
    "HospitalDailyCashRider": {
        "Indicator": "0",
        "SumInsured": 0,
    },
    "OverseasTravelSecure": {
        "Indicator": "0",
    },
    "AggregateDeductible": {
        "Indicator": "0",
    },
    "ProtectBenefit": {
        "Indicator": "0",
    },
    "PlusBenefit": {
        "Indicator": "0",
    },
    "PEDWaiting": {
        "Indicator": "0",
        "PeriodDuration": "0",
    },
    "Parenthood": {
        "Indicator": "0",
        "SumInsured": 0,
    },
    "Limitless": {
        "Indicator": "0",
    },
    "ABCD": {
        "Indicator": "0",
    },
}

FIRST_NAMES = {
    'M':['Aarav','Vihaan','Arjun','Rohan','Kabir','Aditya','Gitesh','Rahul'],
    'F':['Anaya','Nisha','Aditi','Meera','Isha','Kavya','Priya','Saanvi']
}
LAST_NAMES = ['Patil','Sharma','Jain','Mehta','Nair','Deshmukh','Kulkarni','Verma']


# Supported family combinations.
# A = Adult, C = Child. Maximum total members is 8.
FAMILY_TYPES = []
for total_members in range(1, 9):
    for adults in range(1, min(4, total_members) + 1):
        children = total_members - adults
        if children == 0:
            FAMILY_TYPES.append(f"{adults}A")
        else:
            FAMILY_TYPES.append(f"{adults}A+{children}C")


def parse_family_type(value):
    """Return (adult_count, child_count) from values such as 2A or 2A+1C."""
    value = str(value or "1A").strip().upper()
    match = re.fullmatch(r"(\d+)A(?:\+(\d+)C)?", value)
    if not match:
        return 1, 0

    adults = max(1, int(match.group(1)))
    children = int(match.group(2) or 0)

    # Keep the application limit at eight total members.
    if adults + children > 8:
        children = max(0, 8 - adults)

    return adults, children


def random_family_from_type(family_type, proposer_gender="M"):
    """
    Generate relationships that exactly match FamilyType.
    Example: 2A+2C -> Self, Wife/Husband, Son/Daughter, Son/Daughter.
    """
    adults, children = parse_family_type(family_type)

    spouse = "Wife" if proposer_gender == "M" else "Husband"
    adult_pool = [spouse, "Father", "Mother"]
    child_pool = ["Son", "Daughter"]

    relationships = ["Self"]

    # Add the requested number of adults.
    remaining_adults = adults - 1
    random.shuffle(adult_pool)
    relationships.extend(adult_pool[:remaining_adults])

    # This should rarely be needed because FamilyType supports up to 4 adults.
    while len(relationships) < adults:
        relationships.append(random.choice(["Father", "Mother"]))

    # Add the requested number of children.
    for _ in range(children):
        relationships.append(random.choice(child_pool))

    return relationships


def dob_from_age(age):
    today = date.today()
    return date(today.year-age, random.randint(1,12), random.randint(1,28)).isoformat()


def random_family(count, proposer_gender='M'):
    count = max(1, min(int(count), 8))
    spouse = 'Wife' if proposer_gender == 'M' else 'Husband'
    pool = [spouse, 'Son', 'Daughter', 'Father', 'Mother']
    result = ['Self']
    while len(result) < count:
        relation = random.choice(pool)
        if relation in ['Wife','Husband','Father','Mother'] and relation in result:
            continue
        result.append(relation)
    return result


def mock_member(sequence, relation, sum_insured='1000000'):
    p = RELATIONS.get(relation, RELATIONS['Self'])
    age = random.randint(*p['age'])
    gender = p['gender']
    return {
        'SequenceID': str(sequence),
        'IsProposerInsured': '1' if sequence == 1 else p['proposer'],
        'Nationality': 'INDI', 'MaritalStatus': 'S', 'OptFor': '',
        'InsuredLevel80DCertificate': False,
        'RelationshipWithProposer': relation,
        'ResidentialStatus': 'ResidentIndian', 'SumInsured': str(sum_insured),
        'Height': f"{random.uniform(*p['height']):.2f}", 'Gender': gender,
        'MemberDOB': dob_from_age(age), 'Salutation': p['salutation'],
        'Weight': f"{random.uniform(*p['weight']):.2f}",
        'InsuredName': f"{random.choice(FIRST_NAMES[gender])} {random.choice(LAST_NAMES)}",
        'MedicalAdversity': '0', 'IsPoliticallyExposedPerson': '0'
    }


def questions():
    return {'MedicalQuestions': {'LifeStyleQ': {'Lifestylehabits':'0'}, 'Primary': [
        {'Id': q, 'Indicator':'0'} for q in ['ADD','SG','MR','IFT','HIP','PRG','DDAC']
    ]}}


def flag(value, default='0'):
    """Normalize Yes/No, true/false and 1/0 values to API string flags."""
    if value is None:
        return default
    return '1' if str(value).strip().lower() in {'1', 'yes', 'true', 'y'} else '0'


def family_type_from_members(risks):
    child_relations = {"Son", "Daughter"}
    children = sum(
        1 for risk in risks
        if risk["member"].get("RelationshipWithProposer") in child_relations
    )
    adults = len(risks) - children
    if children:
        return f"{adults}A+{children}C"
    return f"{adults}A"


def build_payload(data):
    proposer_gender = data.get("gender", "M")
    policy_type = data.get("policy_type", "Floater")
    selected_plan = data.get("plan", "OptimaSecure")

    # PolicyType and FamilyType are independent.
    # Both Individual and Floater may contain multiple members.
    family_type = data.get("family_type", "1A")
    adult_count, child_count = parse_family_type(family_type)
    member_count = adult_count + child_count

    relationships = (
        data.get("relationships")
        or random_family_from_type(family_type, proposer_gender)
    )
    selected = set(data.get("coverages", []))
    member_overrides = data.get("members", [])

    loyalty_discount = flag(data.get("loyalty_discount"))
    nri_discount = flag(data.get("nri_discount"))
    aggregate_deductible_selection = str(
        data.get("aggregate_deductible", "NO")
    ).strip().upper()

    if aggregate_deductible_selection == "NO":
        aggregate_deductible_discount = "0"
        aggregate_deductible_value = None
    else:
        aggregate_deductible_discount = "1"
        aggregate_deductible_value = aggregate_deductible_selection
    policy_term = str(data.get("term", "12"))
    residential_status = "NRI" if nri_discount == "1" else "ResidentIndian"
    current_country = data.get("current_country_of_residence", "")
    if nri_discount == "1" and not current_country:
        current_country = "United Arab Emirates"

    account_name = data.get("name", "SAKUNTHALA P").strip()
    account_dob = data.get("dob", "1987-09-07")
    account_gender = proposer_gender
    account_salutation = data.get(
        "salutation",
        "MR" if account_gender == "M" else "MRS"
    )

    risks = []
    for i in range(member_count):
        relation = relationships[i] if i < len(relationships) else "Other"
        member = mock_member(i + 1, relation, data.get("sum_insured", "1000000"))

        if i < len(member_overrides):
            member.update({
                k: v for k, v in member_overrides[i].items()
                if v not in (None, "")
            })

        # Always use the main Sum Insured dropdown for every member.
        member["SumInsured"] = str(
            data.get("sum_insured", "1000000")
        )

        # Proposer member must always match account-level identity.
        if i == 0:
            member.update({
                "IsProposerInsured": "1",
                "RelationshipWithProposer": "Self",
                "InsuredName": account_name.replace(" ", ""),
                "MemberDOB": account_dob,
                "Gender": account_gender,
                "Salutation": account_salutation,
            })
        else:
            member["IsProposerInsured"] = "0"

        member["ResidentialStatus"] = residential_status

        coverage_list = []

        critical_illness_enabled = flag(
            data.get("critical_illness_enabled")
        )
        hospital_cash_enabled = flag(
            data.get("hospital_cash_enabled")
        )
        global_plans = {
            "OptimaSecureGlobal",
            "OptimaSecureGlobalPlus",
        }

        for coverage_name, defaults in COVERAGES.items():
            item = {"Type": coverage_name, **defaults}

            if coverage_name == "MyHealthCriticalIllness":
                item["Indicator"] = critical_illness_enabled
                item["PlanName"] = str(
                    data.get("critical_illness_plan", "7")
                )
                item["SumInsured"] = (
                    int(data.get("critical_illness_sum_insured", 100000))
                    if critical_illness_enabled == "1"
                    else 0
                )

            elif coverage_name == "MyHealthHospitalCashBenefit":
                item["Indicator"] = hospital_cash_enabled
                item["SumInsured"] = (
                    int(data.get("hospital_cash_sum_insured", 1000))
                    if hospital_cash_enabled == "1"
                    else 0
                )
                global_requested = flag(
                    data.get("hospital_cash_global")
                )
                item["IsGlobal"] = (
                    "1"
                    if (
                        hospital_cash_enabled == "1"
                        and selected_plan in global_plans
                        and global_requested == "1"
                    )
                    else "0"
                )

            elif coverage_name == "AggregateDeductible":
                # Keep Aggregate Deductible controlled only by its
                # separate dropdown. It is hidden from generic checkboxes.
                item["Indicator"] = aggregate_deductible_discount
                if aggregate_deductible_discount == "1":
                    item["DeductibleValue"] = aggregate_deductible_value
                else:
                    item.pop("DeductibleValue", None)

            else:
                item["Indicator"] = (
                    "1" if coverage_name in selected else "0"
                )

            coverage_list.append(item)

        risk_item = {
            "member": {
                key: value for key, value in member.items()
                if key != "SequenceID"
            },
            "Questions": questions(),
            "coverage": coverage_list,
            "SequenceID": str(i + 1),
        }

        if nri_discount == "1":
            risk_item["nriquestions"] = {
                "PassportNo": data.get(
                    "passport_no",
                    f"P{random.randint(1000000, 9999999)}"
                ),
                "DateOfIssue": data.get(
                    "passport_issue_date",
                    "2024-01-01"
                ),
                "DateOfExpiry": data.get(
                    "passport_expiry_date",
                    "2034-01-01"
                ),
                "VisaNo": data.get(
                    "visa_no",
                    f"V{random.randint(100000, 999999)}"
                ),
                "TelephoneNo": data.get(
                    "nri_telephone_no",
                    data.get("mobile", "9999999999")
                ),
                "Work": {
                    "ResidentialVisaNo": data.get(
                        "residential_visa_no",
                        f"RV{random.randint(100000, 999999)}"
                    )
                },
                "NRIStatus": "NRI",
                "CurrentCountryOfResidence": (
                    current_country or "United Arab Emirates"
                ),
                "NRIDiscount": "1",
                "DateOfReturn": data.get("date_of_return", ""),
            }

        risks.append(risk_item)

    address1 = data.get("address1", "Houseno.61")
    address2 = data.get("address2", "MotheNagaon")
    address3 = data.get("address3", "NearGaneshMandir")
    city = data.get("city", "Mumbai")
    state = data.get("state", "MHIN")
    pincode = data.get("pincode", "400010")
    full_address = data.get(
        "address",
        f"{address1},{address2},{address3},{city},{pincode},Maharashtra"
    )

    normalized_name = account_name.replace(" ", "")
    first_name = data.get(
        "first_name",
        account_name.split()[0] if account_name else "SAKUNTHALA"
    )
    short_name = data.get(
        "short_name",
        account_name.split()[-1][0] if len(account_name.split()) > 1 else "P"
    )

    payload = {
        "session": {
            "data": {
                "account": {
                    "DOBAsPerPehchaan": account_dob,
                    "communication": {
                        "EmailId": data.get("email", "ub@hdfcergo.com"),
                        "Mobile": data.get("mobile", "9930247107"),
                    },
                    "intermediarydetails": {
                        "SP_PospCode": data.get("sp_posp_code")
                    },
                    "nominee": {
                        "AddressSameAsProposer": data.get(
                            "nominee_address_same", "Yes"
                        ),
                        "NomineeName": data.get("nominee_name", "Nisha Patil"),
                        "Address": data.get("nominee_address", full_address),
                        "NomineeDOB": data.get("nominee_dob", "1989-06-24"),
                        "NomineeAge": data.get("nominee_age", "36.00"),
                        "NomineeRelationship": data.get(
                            "nominee_relationship", "Wife"
                        ),
                    },
                    "Salutation": account_salutation,
                    "NameAsPerPehchaanID": data.get(
                        "name_as_per_pehchaan", normalized_name
                    ),
                    "OtherIndustryType": data.get("other_industry_type", ""),
                    "DateOfBirth": account_dob,
                    "PehchaanIDStatus": data.get(
                        "pehchaan_status", "approved"
                    ),
                    "SourceOfIncome": data.get("source_of_income", ""),
                    "MaritalStatus": data.get("marital_status", "S"),
                    "LimitMoreThanFiveLakhs": data.get(
                        "limit_more_than_five_lakhs", "0"
                    ),
                    "SourceOfFunds": data.get("source_of_funds"),
                    "PhysicalPolicyCopyrequired": data.get(
                        "physical_policy_copy_required", "0"
                    ),
                    "CISAcknowledgement": data.get(
                        "cis_acknowledgement", "No"
                    ),
                    "Nationality": data.get("nationality", "INDI"),
                    "Gender": account_gender,
                    "PoliticallyExposed": data.get(
                        "politically_exposed", "0"
                    ),
                    "NRIDiscount": nri_discount,
                    "Name": short_name,
                    "kyc": {
                        "IsKYCVerified": data.get("kyc_verified", "1")
                    },
                    "CKYCNumber": data.get("ckyc_number", ""),
                    "AnnualIncome": data.get(
                        "annual_income", "15-20Lacs"
                    ),
                    "refunddetails": {
                        "BankAccountNumber": data.get(
                            "bank_account", "178293738399"
                        ),
                        "BranchName": data.get(
                            "branch", "ERNAKULAMCOCHIN"
                        ),
                        "ChequeAmount": data.get("cheque_amount"),
                        "NameAsInBankAccount": data.get(
                            "bank_account_name", normalized_name
                        ),
                        "IFSCCode": data.get("ifsc", "ICIC0000010"),
                        "BankName": data.get(
                            "bank_name", "ICICIBANKLIMITED"
                        ),
                    },
                    "CriminalProceedingDetails": data.get(
                        "criminal_proceeding_details"
                    ),
                    "LimitMoreThanTwoLakhs": data.get(
                        "limit_more_than_two_lakhs", "0"
                    ),
                    "OtherEducation": data.get("other_education", ""),
                    "CriminalProceedingInPast3Years": data.get(
                        "criminal_proceeding_past_3_years", "0"
                    ),
                    "address": {
                        "Permanent": {
                            "Address3": address3,
                            "Address2": address2,
                            "State": state,
                            "City": city,
                            "Pincode": pincode,
                            "Address1": address1,
                        },
                        "Correspondence": {
                            "City": city,
                            "Address1": data.get(
                                "correspondence_address1", "Houseno. 61"
                            ),
                            "Address3": data.get(
                                "correspondence_address3",
                                "Near Ganesh Mandir"
                            ),
                            "State": state,
                            "Pincode": pincode,
                            "Address2": data.get(
                                "correspondence_address2", "Mothe Nagaon"
                            ),
                        },
                    },
                    "IndustryType": data.get("industry_type", ""),
                    "FirstName": first_name,
                    "PehchaanID": data.get(
                        "pehchaan_id", "UC4VNKAALM"
                    ),
                    "PartyType": data.get("party_type", "P"),
                    "TypeOfBusiness": data.get(
                        "type_of_business", "Intermediary"
                    ),
                    "PANNumber": data.get("pan", "CYUPS5288Q"),
                    "ResidentialStatus": residential_status,
                    "MiddleName": data.get("middle_name", ""),
                    "InvestableAssets": data.get(
                        "investable_assets", "0"
                    ),
                    "Education": data.get(
                        "education", "PostGraduate"
                    ),
                    "Occupation": data.get("occupation", "231"),
                    "IsPermanentAddressSameAsCorrespondence": data.get(
                        "same_address", "0"
                    ),
                },
                "policy": {
                    "line": {"risk": risks},
                    "ProposalDate": data.get("proposal_date"),
                    "PolicyType": policy_type,
                    "Plan": selected_plan,
                    "bancassurance": {
                        "LosNoLanAan": data.get("los_no_lan_aan"),
                        "BankBranchID": data.get("bank_branch_id"),
                        "SavingBankAccount": data.get(
                            "saving_bank_account"
                        ),
                        "LCCode": data.get("lc_code"),
                        "Funded": data.get("funded", "0"),
                        "CustomerID": data.get("customer_id"),
                        "LGCode": data.get("lg_code"),
                        "LOSCode": data.get("los_code"),
                    },
                    "EffectiveTime": data.get(
                        "effective_time", "03:30PM"
                    ),
                    "Product": selected_plan,
                    "Term": policy_term,
                    "EMI": data.get("emi", "0"),
                    "ChildSplitFlag": data.get(
                        "child_split_flag", "0"
                    ),
                    "UserID": data.get(
                        "user_id", "AKSHAY.JAINPOS"
                    ),
                    "InstallmentFrequencyOpted": data.get(
                        "installment_frequency", "HEGI_ACT_P1"
                    ),
                    "EffectiveDate": data.get("effective_date"),
                    "FamilyType": family_type,
                    "OriginatingSystem": data.get(
                        "originating_system", "1UP"
                    ),
                    "SuppressEmailSMS": data.get(
                        "suppress_email_sms", "1"
                    ),
                    "TransactionType": data.get(
                        "transaction_type", "New"
                    ),
                    "IsEmployeeDiscountApplicable": data.get(
                        "employee_discount", "0"
                    ),
                    "IsProposal": data.get("is_proposal", "0"),
                    "IsLoyaltyDiscountApplicable": loyalty_discount,
                },
                "InstallmentSchedule": {
                    "ProvidePayPlan": data.get(
                        "provide_pay_plan", "1"
                    )
                },
                "newquoteselection": {
                    "Vertical": data.get(
                        "vertical", "101011000"
                    ),
                    "Intermediary": data.get(
                        "intermediary", "200798503757"
                    ),
                },
            },
            "@clientId": data.get("client_id", ""),
            "@id": data.get("id", ""),
        }
    }

    if nri_discount == "1":
        payload["session"]["data"]["account"][
            "CurrentCountryOfResidence"
        ] = current_country

    return payload


def _json_or_text(response):
    try:
        return response.json()
    except ValueError:
        return {"rawText": response.text}


def _api_error(step, response):
    body = _json_or_text(response)
    raise RuntimeError(
        f"{step} failed with HTTP {response.status_code}: "
        f"{json.dumps(body, ensure_ascii=False)[:1500]}"
    )


def get_access_token(force_refresh=False):
    now = time.time()
    if (
        not force_refresh
        and _TOKEN_CACHE["access_token"]
        and now < _TOKEN_CACHE["expires_at"] - 60
    ):
        return _TOKEN_CACHE["access_token"]

    if not API_CLIENT_ID or not API_CLIENT_SECRET:
        raise RuntimeError(
            "API credentials are missing. Set API_CLIENT_ID and "
            "API_CLIENT_SECRET in your local .env file."
        )

    response = requests.post(
        AUTH_URL,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data={
            "grant_type": "client_credentials",
            "client_id": API_CLIENT_ID,
            "client_secret": API_CLIENT_SECRET,
        },
        timeout=API_TIMEOUT_SECONDS,
    )
    if not response.ok:
        _api_error("Authentication", response)

    data = _json_or_text(response)
    token = (
        data.get("access_token")
        or data.get("accessToken")
        or data.get("token")
        or data.get("authToken")
    )
    if not token:
        raise RuntimeError(
            "Authentication succeeded but no access token was found in the response."
        )

    expires_in = int(data.get("expires_in") or data.get("expiresIn") or 3600)
    _TOKEN_CACHE["access_token"] = token
    _TOKEN_CACHE["expires_at"] = now + expires_in
    return token


def api_headers(token, include_cookie=False):
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}",
    }
    if include_cookie and PAYMENT_COOKIE:
        headers["Cookie"] = PAYMENT_COOKIE
    return headers


def post_api(step, url, payload, token, include_cookie=False):
    response = requests.post(
        url,
        headers=api_headers(token, include_cookie=include_cookie),
        json=payload,
        timeout=API_TIMEOUT_SECONDS,
    )
    if response.status_code == 401:
        token = get_access_token(force_refresh=True)
        response = requests.post(
            url,
            headers=api_headers(token, include_cookie=include_cookie),
            json=payload,
            timeout=API_TIMEOUT_SECONDS,
        )
    if not response.ok:
        _api_error(step, response)
    return _json_or_text(response), token


def find_first(obj, candidate_keys):
    normalized = {re.sub(r"[^a-z0-9]", "", key.lower()) for key in candidate_keys}
    if isinstance(obj, dict):
        for key, value in obj.items():
            key_norm = re.sub(r"[^a-z0-9]", "", str(key).lower())
            if key_norm in normalized and value not in (None, "", [], {}):
                return value
        for value in obj.values():
            found = find_first(value, candidate_keys)
            if found not in (None, "", [], {}):
                return found
    elif isinstance(obj, list):
        for value in obj:
            found = find_first(value, candidate_keys)
            if found not in (None, "", [], {}):
                return found
    return None


def get_nested(obj, path):
    """Read an exact nested response path, including list indexes."""
    current = obj
    for part in path:
        if isinstance(part, int):
            if not isinstance(current, list) or len(current) <= part:
                return None
            current = current[part]
        else:
            if not isinstance(current, dict) or part not in current:
                return None
            current = current[part]
    return current


def require_value(response, keys, label, step):
    value = find_first(response, keys)
    if value in (None, "", [], {}):
        raise RuntimeError(
            f"{step} response did not contain {label}. "
            f"Checked keys: {', '.join(keys)}"
        )
    return value


def build_payment_payload(quote_response, options):
    quote_id = require_value(
        quote_response,
        ["QuoteId", "QuoteID", "ProposalId", "ProposalID"],
        "QuoteId/ProposalId",
        "Create Quote",
    )
    quote_number = require_value(
        quote_response,
        ["QuoteNumber", "QuoteNo", "ProposalNumber", "ProposalNo"],
        "QuoteNumber",
        "Create Quote",
    )
    # Postman collection maps PayorPartyId from {{ClientID}}.
    # Check ClientID first, then preserve older fallback aliases.
    payor_party_id = require_value(
        quote_response,
        [
            "ClientID", "ClientId", "clientId", "clientID",
            "PayorPartyId", "PayerPartyId", "PartyId", "PartyID",
            "CustomerPartyId", "CustomerPartyID"
        ],
        "ClientID/PayorPartyId",
        "Create Quote",
    )
    amount = (
        find_first(
            quote_response,
            [
                "Instrumentamount", "InstrumentAmount", "TotalPremium",
                "GrossPremium", "FinalPremium", "PremiumAmount", "Amount"
            ],
        )
        or options.get("instrument_amount")
    )
    if amount in (None, ""):
        raise RuntimeError(
            "Create Quote response did not contain a premium amount and no manual "
            "payment amount was supplied."
        )

    return {
        "PaymentRequest": {
            "CreatePaymentRequest": [{
                "PayorPartyId": str(payor_party_id),
                "PaymentMethodCode": options.get("payment_method_code", "EP"),
                "InstrumentNumber": options.get("instrument_number", f"AUTO{int(time.time())}"),
                "InstrumentDate": options.get("instrument_date"),
                "Instrumentamount": str(amount),
                "PaymentSequeceId": "1",
                "PGName": options.get("pg_name", "Razorpay"),
                "MandateId": options.get("mandate_id", ""),
                "EntryUserId": options.get("entry_user_id", "quote.generator"),
            }],
            "Lob": "Health",
            "QuoteId": str(quote_id),
            "QuoteNumber": str(quote_number),
            "SourceSystem": "PORTAL",
            "Remarks": str(quote_number),
            "PaymentCollected": 1,
        }
    }


def build_issue_payload(quote_response, payment_response):
    proposal_id = (
        find_first(payment_response, ["ProposalId", "ProposalID", "QuoteId", "QuoteID"])
        or find_first(quote_response, ["ProposalId", "ProposalID", "QuoteId", "QuoteID"])
    )
    if proposal_id in (None, ""):
        raise RuntimeError("Unable to find ProposalId for issuePolicy.")

    # Exact Postman mapping:
    # createPaymentResponse[0].session.data.CreatePaymentResponse
    # .PaymentDetail.PaymentAllocations.PaymentAllocation.DestinationId
    suspense_id = get_nested(
        payment_response,
        [
            "createPaymentResponse", 0, "session", "data",
            "CreatePaymentResponse", "PaymentDetail",
            "PaymentAllocations", "PaymentAllocation", "DestinationId"
        ],
    )

    # Keep recursive fallbacks in case the response envelope changes slightly.
    if suspense_id in (None, "", [], {}):
        suspense_id = require_value(
            payment_response,
            [
                "DestinationId", "DestinationID",
                "SuspenseId", "SuspenseID",
                "SuspenseNumber", "SuspenseNo"
            ],
            "DestinationId/SuspenseId",
            "Create Payment",
        )
    if not isinstance(suspense_id, list):
        suspense_id = [suspense_id]

    return {
        "IssuePolicyRequest": {
            "ProposalId": int(proposal_id) if str(proposal_id).isdigit() else proposal_id,
            "SuspenseList": {"SuspenseId": suspense_id},
            "SuppressEmailSMS": "0",
        }
    }



def normalize_numeric(value):
    """Convert numeric strings to int, preserving non-numeric IDs."""
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.isdigit():
            return int(stripped)
    return value

def execute_policy_workflow(quote_payload, payment_input=None):
    payment_input = payment_input or {}

    token = get_access_token()

    # 1. CREATE QUOTE
    risks_to_send = get_nested(
        quote_payload,
        ["session", "data", "policy", "line", "risk"],
    )

    if not isinstance(risks_to_send, list) or not risks_to_send:
        raise RuntimeError(
            "Create Quote payload does not contain "
            "session.data.policy.line.risk members."
        )

    try:
        with open(
            "actual_create_quote_request.json",
            "w",
            encoding="utf-8",
        ) as request_file:
            json.dump(
                quote_payload,
                request_file,
                indent=2,
                ensure_ascii=False,
            )
    except OSError:
        app.logger.warning(
            "Could not save actual_create_quote_request.json"
        )

    app.logger.info(
        "Create Quote sending %s risk object(s).",
        len(risks_to_send),
    )

    quote_response, token = post_api(
        "Create Quote",
        CREATE_QUOTE_URL,
        quote_payload,
        token,
    )

    try:
        with open(
            "actual_create_quote_response.json",
            "w",
            encoding="utf-8",
        ) as response_file:
            json.dump(
                quote_response,
                response_file,
                indent=2,
                ensure_ascii=False,
            )
    except OSError:
        app.logger.warning(
            "Could not save actual_create_quote_response.json"
        )

    # Exact Create Quote mappings from the Postman collection:
    # QuoteID        = response.session["@id"]
    # ClientID       = response.session["@clientId"]
    # QuoteNumber    = response.session.data.policy.QuoteNumber
    # PremiumAmount  = response.session.data.policy.line.Premium
    # InstrumentDate = response.session.data.policy.EffectiveDate

    quote_id = get_nested(
        quote_response,
        ["session", "@id"],
    )
    client_id = get_nested(
        quote_response,
        ["session", "@clientId"],
    )
    quote_number = get_nested(
        quote_response,
        ["session", "data", "policy", "QuoteNumber"],
    )
    premium_amount = get_nested(
        quote_response,
        ["session", "data", "policy", "line", "Premium"],
    )
    instrument_date = get_nested(
        quote_response,
        ["session", "data", "policy", "EffectiveDate"],
    )

    missing_paths = []

    if quote_id in (None, "", [], {}):
        missing_paths.append('session["@id"]')

    if client_id in (None, "", [], {}):
        missing_paths.append('session["@clientId"]')

    if quote_number in (None, "", [], {}):
        missing_paths.append(
            "session.data.policy.QuoteNumber"
        )

    if premium_amount in (None, "", [], {}):
        missing_paths.append(
            "session.data.policy.line.Premium"
        )

    if instrument_date in (None, "", [], {}):
        missing_paths.append(
            "session.data.policy.EffectiveDate"
        )

    if missing_paths:
        raise RuntimeError(
            "Create Quote response is missing required Postman mapping path(s): "
            + ", ".join(missing_paths)
        )

    payment_payload = {
        "PaymentRequest": {
            "CreatePaymentRequest": [
                {
                    "PayorPartyId": str(client_id),
                    "PaymentMethodCode": payment_input.get(
                        "payment_method_code", "EP"
                    ),
                    "InstrumentNumber": payment_input.get(
                        "instrument_number"
                    ) or f"AUTO{int(time.time())}",
                    # Exact Postman mapping:
                    # response.session.data.policy.EffectiveDate
                    "InstrumentDate": str(instrument_date),

                    # Exact Postman mapping:
                    # response.session.data.policy.line.Premium
                    "Instrumentamount": str(premium_amount),

                    "PaymentSequeceId": "1",
                    "PGName": "Razorpay",
                    "MandateId": "23232332332",
                    "EntryUserId": "umesh.b",
                }
            ],
            "Lob": "Health",
            "QuoteId": str(quote_id),
            "QuoteNumber": str(quote_number),
            "SourceSystem": "PORTAL",
            "Remarks": str(quote_number),
            "PaymentCollected": 1,
        }
    }

    # 2. CREATE PAYMENT
    try:
        with open(
            "actual_create_payment_request.json",
            "w",
            encoding="utf-8",
        ) as payment_request_file:
            json.dump(
                payment_payload,
                payment_request_file,
                indent=2,
                ensure_ascii=False,
            )
    except OSError:
        app.logger.warning(
            "Could not save actual_create_payment_request.json"
        )

    payment_response, token = post_api(
        "Create Payment",
        CREATE_PAYMENT_URL,
        payment_payload,
        token,
        include_cookie=True,
    )

    try:
        with open(
            "actual_create_payment_response.json",
            "w",
            encoding="utf-8",
        ) as payment_response_file:
            json.dump(
                payment_response,
                payment_response_file,
                indent=2,
                ensure_ascii=False,
            )
    except OSError:
        app.logger.warning(
            "Could not save actual_create_payment_response.json"
        )

    # Exact Postman mapping.
    suspense_id = get_nested(
        payment_response,
        [
            "createPaymentResponse", 0, "session", "data",
            "CreatePaymentResponse", "PaymentDetail",
            "PaymentAllocations", "PaymentAllocation",
            "DestinationId",
        ],
    )

    if suspense_id in (None, "", [], {}):
        suspense_id = require_value(
            payment_response,
            ["DestinationId", "SuspenseId", "SuspenseID"],
            "DestinationId/SuspenseID",
            "Create Payment",
        )

    if isinstance(suspense_id, list):
        suspense_ids = [
            normalize_numeric(item)
            for item in suspense_id
        ]
    else:
        suspense_ids = [
            normalize_numeric(suspense_id)
        ]

    normalized_quote_id = normalize_numeric(quote_id)

    # 3. ISSUE POLICY
    # Exact Postman payload: numeric QuoteID and numeric SuspenseID.
    issue_payload = {
        "IssuePolicyRequest": {
            "ProposalId": normalized_quote_id,
            "SuspenseList": {
                "SuspenseId": suspense_ids
            },
            "SuppressEmailSMS": "0"
        }
    }

    # Payment allocation may take a few seconds to become available
    # to Issue Policy. Postman naturally has a delay between requests,
    # so reproduce that behaviour and retry safely.
    time.sleep(ISSUE_POLICY_INITIAL_DELAY)

    issue_response = None
    raw_policy_number = None
    issue_attempts = []

    for attempt in range(1, ISSUE_POLICY_MAX_ATTEMPTS + 1):
        issue_response, token = post_api(
            f"Issue Policy (attempt {attempt})",
            ISSUE_POLICY_URL,
            issue_payload,
            token,
        )

        raw_policy_number = get_nested(
            issue_response,
            ["session", "data", "policy", "PolicyNumber"],
        )

        issue_attempts.append({
            "attempt": attempt,
            "policyNumber": raw_policy_number,
            "response": issue_response,
        })

        if raw_policy_number not in (None, "", [], {}, "0", 0):
            break

        if attempt < ISSUE_POLICY_MAX_ATTEMPTS:
            time.sleep(ISSUE_POLICY_RETRY_DELAY)

    if raw_policy_number in (None, "", [], {}, "0", 0):
        raise RuntimeError(
            "Issue Policy did not return a valid "
            "session.data.policy.PolicyNumber after "
            f"{ISSUE_POLICY_MAX_ATTEMPTS} attempts.\n\n"
            "Issue Policy request sent:\n"
            + json.dumps(issue_payload, indent=2, ensure_ascii=False)
            + "\n\nAll Issue Policy attempts:\n"
            + json.dumps(issue_attempts, indent=2, ensure_ascii=False)
        )

    sixteen_digit_policy_number = f"{raw_policy_number}00"

    # 4. RENEWAL BASE QUOTE
    renewal_payload = {
        "policyNumber": sixteen_digit_policy_number,
        "forceCreateQuote": True,
    }

    renewal_response, token = post_api(
        "Renewal Base Quote",
        RENEWAL_BASE_QUOTE_URL,
        renewal_payload,
        token,
    )

    renewal_quote_number = find_first(
        renewal_response,
        ["QuoteNumber", "RenewalQuoteNumber", "QuoteNo"],
    )

    return {
        "success": True,
        "message": "Renewal policy created successfully.",
        "summary": {
            "QuoteID": normalized_quote_id,
            "ClientID": client_id,
            "QuoteNumber": quote_number,
            "PremiumAmount": premium_amount,
            "InstrumentDate": instrument_date,
            "CreatePaymentMapping": {
                "PayorPartyId": str(client_id),
                "QuoteId": str(quote_id),
                "QuoteNumber": str(quote_number),
                "PremiumAmount": str(premium_amount),
                "InstrumentDate": str(instrument_date),
            },
            "RiskCountSent": len(risks_to_send),
            "SuspenseID": suspense_ids,
            "rawPolicyNumber": raw_policy_number,
            "policyNumber": sixteen_digit_policy_number,
            "renewalQuoteNumber": renewal_quote_number,
        },
        "requests": {
            "createQuote": quote_payload,
            "createPayment": payment_payload,
            "issuePolicy": issue_payload,
            "renewalBaseQuote": renewal_payload,
        },
        "responses": {
            "createQuote": quote_response,
            "createPayment": payment_response,
            "issuePolicy": issue_response,
            "issuePolicyAttempts": issue_attempts,
            "renewalBaseQuote": renewal_response,
        },
    }


@app.post('/api/run-policy-workflow')
def api_run_policy_workflow():
    try:
        body = request.get_json(silent=True)

        if not isinstance(body, dict):
            return jsonify({
                "success": False,
                "error": "Request body must be valid JSON."
            }), 400

        quote_payload = body.get("quote_payload")
        payment_input = body.get("payment") or {}

        if not isinstance(quote_payload, dict):
            return jsonify({
                "success": False,
                "error": "quote_payload is missing or invalid."
            }), 400

        result = execute_policy_workflow(
            quote_payload,
            payment_input,
        )
        return jsonify(result), 200

    except requests.Timeout:
        app.logger.exception("Policy workflow API timeout")
        return jsonify({
            "success": False,
            "error": "External API timed out. Please try again."
        }), 504

    except requests.RequestException as exc:
        app.logger.exception("Policy workflow network error")
        return jsonify({
            "success": False,
            "error": f"External API network error: {exc}"
        }), 502

    except Exception as exc:
        app.logger.exception("Policy workflow failed")
        return jsonify({
            "success": False,
            "error": str(exc)
        }), 500



# ---------------------------------------------------------------------
# BULK EXCEL -> JSON ZIP
# This feature is independent of the existing single-scenario workflow.
# ---------------------------------------------------------------------

def _bulk_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _bulk_yes(value):
    return _bulk_text(value).lower() in {
        "yes", "y", "true", "1", "selected"
    }


def _bulk_number(value, default=0):
    """Convert Excel values such as 5,00,000 / 50 K / 2 Lac to integers."""
    if value in (None, ""):
        return default

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        return int(value)

    text = _bulk_text(value).lower().replace(",", "").replace("₹", "")
    text = text.replace("rupees", "").strip()

    multiplier = 1

    if "crore" in text or re.search(r"\bcr\b", text):
        multiplier = 10000000
    elif "lakh" in text or "lac" in text:
        multiplier = 100000
    elif re.search(r"\bk\b", text):
        multiplier = 1000

    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return default

    return int(float(match.group()) * multiplier)


def _bulk_date(value):
    """Return Excel date/datetime/string as YYYY-MM-DD."""
    if value in (None, ""):
        return ""

    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")

    text = _bulk_text(value)

    for fmt in (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    return text


def _bulk_plan(value):
    text = re.sub(r"[^a-z0-9]", "", _bulk_text(value).lower())

    mapping = {
        "optimasecure": "OptimaSecure",
        "optimasecureglobal": "OptimaSecureGlobal",
        "optimasecureglobalplus": "OptimaSecureGlobalPlus",
        "optimalite": "OptimaLite",
    }

    return mapping.get(text, "OptimaSecure")


def _bulk_policy_type(value):
    text = _bulk_text(value).lower()
    return "Floater" if "floater" in text else "Individual"


def _bulk_family_type(value):
    """
    Convert Excel family types such as 2A+1K / 1A+6K to app values
    such as 2A+1C / 1A+6C.
    """
    text = _bulk_text(value).upper().replace(" ", "")
    text = text.replace("K", "C")

    match = re.search(r"(\d+)A(?:\+(\d+)C)?", text)
    if not match:
        return "1A"

    adults = max(1, int(match.group(1)))
    children = int(match.group(2) or 0)

    if children:
        return f"{adults}A+{children}C"
    return f"{adults}A"


def _bulk_tenure_months(value):
    years = _bulk_number(value, 1)

    # Some sheets may already provide 12/24/36.
    if years in {12, 24, 36}:
        return str(years)

    years = max(1, min(years, 3))
    return str(years * 12)


def _bulk_payment_frequency(value):
    text = _bulk_text(value).lower().replace("-", " ")

    if "month" in text:
        return "12"
    if "quarter" in text:
        return "4"
    if "half" in text:
        return "2"

    # Single / Full Payment
    return "0"


def _bulk_ci_plan(value):
    match = re.search(r"([1-7])", _bulk_text(value))
    return match.group(1) if match else "7"


def _bulk_relationship_profile(relation):
    normalized = _bulk_text(relation).lower()

    female_words = (
        "wife", "spouse", "spouce", "mother", "daughter",
        "sister", "female"
    )
    male_words = (
        "husband", "father", "son", "brother", "male", "self"
    )

    if any(word in normalized for word in female_words):
        return "F", "MRS" if "mother" in normalized or "wife" in normalized else "MISS"

    if any(word in normalized for word in male_words):
        return "M", "MASTER" if "son" in normalized else "MR"

    return "M", "MR"


def _bulk_clean_relation(value):
    text = _bulk_text(value).lower()

    if "self" in text:
        return "Self"
    if "wife" in text or "spouse" in text or "spouce" in text:
        return "Wife"
    if "husband" in text:
        return "Husband"
    if "father" in text:
        return "Father"
    if "mother" in text:
        return "Mother"
    if "daughter" in text:
        return "Daughter"
    if "son" in text:
        return "Son"
    if "sister in law" in text:
        return "SisterInLaw"
    if "brother in law" in text:
        return "BrotherInLaw"
    if "sister" in text:
        return "Sister"
    if "brother" in text:
        return "Brother"

    return _bulk_text(value).title() or "Other"


def _bulk_relationships(value, family_type):
    """
    Build the exact number of relations required by FamilyType.
    Handles phrases such as:
    Self, spouse, son
    Father, mother + 4CD
    Self + 3CD
    """
    adults, children = parse_family_type(family_type)
    total = adults + children

    raw = _bulk_text(value)
    lowered = raw.lower()

    relations = []

    # Remove child-count shorthand before splitting.
    child_count_match = re.search(
        r"(\d+)\s*(?:cd|child|children|kids?|k)\b",
        lowered,
    )
    shorthand_children = (
        int(child_count_match.group(1))
        if child_count_match
        else 0
    )

    cleaned = re.sub(
        r"\+?\s*\d+\s*(?:cd|child|children|kids?|k)\b",
        "",
        raw,
        flags=re.I,
    )

    parts = [
        part.strip()
        for part in re.split(r"[,;+]", cleaned)
        if part.strip()
    ]

    for part in parts:
        relation = _bulk_clean_relation(part)
        if relation:
            relations.append(relation)

    # Guarantee proposer first when Excel does not explicitly say Self.
    if "Self" in relations:
        relations.remove("Self")
    relations.insert(0, "Self")

    current_children = sum(
        1 for relation in relations
        if relation in {"Son", "Daughter"}
    )

    required_children = max(children, shorthand_children)

    while current_children < required_children:
        relations.append(
            "Son" if current_children % 2 == 0 else "Daughter"
        )
        current_children += 1

    # Fill missing adults using spouse/parents.
    adult_fill = ["Wife", "Father", "Mother", "Brother", "Sister"]
    while len([
        r for r in relations
        if r not in {"Son", "Daughter"}
    ]) < adults:
        for candidate in adult_fill:
            if candidate not in relations:
                relations.append(candidate)
                break
        else:
            relations.append("Other")

    # If still short, fill remaining positions with children.
    while len(relations) < total:
        relations.append(
            "Son" if len(relations) % 2 == 0 else "Daughter"
        )

    return relations[:total]


def _bulk_ages(value, member_count):
    ages = [
        int(number)
        for number in re.findall(r"\d+", _bulk_text(value))
    ]

    if not ages:
        ages = [35]

    while len(ages) < member_count:
        ages.append(10 if len(ages) else 35)

    return ages[:member_count]


def _bulk_dob_from_age(age, reference_date):
    try:
        reference = datetime.strptime(reference_date, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        reference = date.today()

    year = max(1900, reference.year - int(age))
    return date(year, 6, 15).isoformat()


def _bulk_member_overrides(
    relationships,
    ages,
    effective_date,
    sum_insured,
):
    members = []

    for index, relation in enumerate(relationships):
        gender, salutation = _bulk_relationship_profile(relation)

        if index == 0:
            gender = "M"
            salutation = "MR"

        members.append({
            "SequenceID": str(index + 1),
            "RelationshipWithProposer": relation,
            "MemberDOB": _bulk_dob_from_age(
                ages[index],
                effective_date,
            ),
            "Gender": gender,
            "Salutation": salutation,
            "SumInsured": str(sum_insured),
            "IsProposerInsured": "1" if index == 0 else "0",
        })

    return members


def _bulk_row_to_input(row):
    """
    Convert one Excel row into the same minimum-input dictionary that
    the existing build_payload() function already accepts.
    """
    family_type = _bulk_family_type(row.get("No of Insureds"))
    relationships = _bulk_relationships(
        row.get("Relationship"),
        family_type,
    )

    effective_date = _bulk_date(row.get("Start Date"))
    proposal_date = effective_date

    sum_insured = _bulk_number(
        row.get("Sum Insured"),
        1000000,
    )

    ages = _bulk_ages(
        row.get("Age"),
        len(relationships),
    )

    selected_coverages = []

    simple_rider_map = {
        "UR Rider": "UnlimitedRestoreRider",
        "OST Rider": "OverseasTravelSecure",
        "Protect Benefit": "ProtectBenefit",
        "Plus Benefit": "PlusBenefit",
        "PED Waiting Period": "PEDWaiting",
        "IPA Rider": "IndividualPersonalAccidentRider",
        "OPD Rider": "OptimaWellbeing",
        "ABCD": "ABCD",
        "Limitless": "Limitless",
        "Parenthood": "Parenthood",
    }

    for excel_column, coverage_type in simple_rider_map.items():
        if _bulk_yes(row.get(excel_column)):
            selected_coverages.append(coverage_type)

    plan = _bulk_plan(row.get("Plan"))

    ci_enabled = _bulk_yes(row.get("CI Rider"))
    hospital_cash_enabled = _bulk_yes(row.get("HDC Rider"))

    hdc_plan = _bulk_text(row.get("HDC Plan")).lower()
    hospital_cash_global = (
        hospital_cash_enabled
        and "global" in hdc_plan
        and plan in {
            "OptimaSecureGlobal",
            "OptimaSecureGlobalPlus",
        }
    )

    deductible_raw = row.get("Deductible")
    deductible_value = _bulk_number(deductible_raw, 0)
    aggregate_deductible = (
        str(deductible_value)
        if deductible_value > 0
        else "NO"
    )

    return {
        "policy_type": _bulk_policy_type(
            row.get("Policy Type")
        ),
        "family_type": family_type,
        "member_count": len(relationships),
        "gender": "M",
        "relationships": relationships,
        "members": _bulk_member_overrides(
            relationships,
            ages,
            effective_date,
            sum_insured,
        ),
        "sum_insured": str(sum_insured),
        "plan": plan,
        "proposal_date": proposal_date,
        "effective_date": effective_date,
        "term": _bulk_tenure_months(
            row.get("Tenure")
        ),
        "emi": _bulk_payment_frequency(
            row.get("Payment Frequency")
        ),
        "aggregate_deductible": aggregate_deductible,
        "employee_discount": flag(
            row.get("Emp Disc")
        ),
        "loyalty_discount": flag(
            row.get("Loyalty Disc")
        ),
        "nri_discount": flag(
            row.get("NRI Disc")
        ),
        "mobile": _bulk_text(
            row.get("Mobile No")
        ),
        "email": _bulk_text(
            row.get("Email ID")
        ),
        "city": _bulk_text(
            row.get("City/District")
        ),
        "pincode": _bulk_text(
            row.get("Pincode")
        ),
        "coverages": selected_coverages,
        "critical_illness_enabled": ci_enabled,
        "critical_illness_plan": _bulk_ci_plan(
            row.get("CI Plan")
        ),
        "critical_illness_sum_insured": str(
            _bulk_number(
                row.get("CI Sum Insured"),
                100000,
            )
        ),
        "hospital_cash_enabled": hospital_cash_enabled,
        "hospital_cash_sum_insured": str(
            _bulk_number(
                row.get("HDC SI"),
                1000,
            )
        ),
        "hospital_cash_global": hospital_cash_global,
    }


def _bulk_safe_filename(value, fallback):
    text = _bulk_text(value) or fallback
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text)
    return text.strip("._") or fallback


@app.post("/api/bulk-generate")
def api_bulk_generate():
    """
    Upload one .xlsx file and return one ZIP containing one JSON
    per non-empty Excel test-case row.
    """
    uploaded_file = request.files.get("excel_file")

    if not uploaded_file or not uploaded_file.filename:
        return jsonify({
            "success": False,
            "error": "Please select an Excel .xlsx file."
        }), 400

    if not uploaded_file.filename.lower().endswith(".xlsx"):
        return jsonify({
            "success": False,
            "error": "Only .xlsx Excel files are supported."
        }), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({
            "success": False,
            "error": (
                "The openpyxl package is missing. "
                "Run: pip install openpyxl"
            )
        }), 500

    try:
        workbook = load_workbook(
            uploaded_file,
            data_only=True,
            read_only=True,
        )

        sheet = (
            workbook["UW"]
            if "UW" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )

        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)

        if not headers:
            raise ValueError("The Excel sheet is empty.")

        normalized_headers = [
            _bulk_text(header)
            for header in headers
        ]

        required_columns = {
            "Test Case S.No.",
            "Policy Type",
            "No of Insureds",
            "Relationship",
            "Sum Insured",
        }

        missing_columns = sorted(
            required_columns.difference(normalized_headers)
        )

        if missing_columns:
            raise ValueError(
                "Missing required Excel column(s): "
                + ", ".join(missing_columns)
            )

        generated_files = []
        summary_rows = []
        filename_counts = {}

        for excel_row_number, values in enumerate(rows, start=2):
            row = dict(zip(normalized_headers, values))

            test_case_number = row.get("Test Case S.No.")

            # Ignore completely empty rows and summary/footer rows.
            if test_case_number in (None, ""):
                continue

            input_data = _bulk_row_to_input(row)
            payload = build_payload(input_data)

            file_stem = _bulk_safe_filename(
                test_case_number,
                f"ROW_{excel_row_number}",
            )

            base_filename = f"TC_{file_stem}"
            filename_counts[base_filename] = (
                filename_counts.get(base_filename, 0) + 1
            )
            occurrence = filename_counts[base_filename]
            filename = (
                f"{base_filename}.json"
                if occurrence == 1
                else f"{base_filename}_{occurrence}.json"
            )
            generated_files.append((
                filename,
                json.dumps(
                    payload,
                    indent=2,
                    ensure_ascii=False,
                ),
            ))

            summary_rows.append({
                "Test Case S.No.": test_case_number,
                "JSON File": filename,
                "Policy Type": input_data["policy_type"],
                "Family Type": input_data["family_type"],
                "Plan": input_data["plan"],
                "Term": input_data["term"],
                "EMI": input_data["emi"],
                "Risk Count": input_data["member_count"],
            })

        if not generated_files:
            raise ValueError(
                "No test-case rows were found in the uploaded Excel file."
            )

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(
            zip_buffer,
            mode="w",
            compression=zipfile.ZIP_DEFLATED,
        ) as archive:
            for filename, content in generated_files:
                archive.writestr(filename, content)

            summary_buffer = io.StringIO()
            writer = csv.DictWriter(
                summary_buffer,
                fieldnames=[
                    "Test Case S.No.",
                    "JSON File",
                    "Policy Type",
                    "Family Type",
                    "Plan",
                    "Term",
                    "EMI",
                    "Risk Count",
                ],
            )
            writer.writeheader()
            writer.writerows(summary_rows)

            archive.writestr(
                "bulk_generation_summary.csv",
                summary_buffer.getvalue(),
            )

        zip_buffer.seek(0)

        response = send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name="bulk_quote_testcases.zip",
        )
        response.headers["X-Generated-Count"] = str(
            len(generated_files)
        )
        return response

    except Exception as exc:
        app.logger.exception("Bulk Excel JSON generation failed")
        return jsonify({
            "success": False,
            "error": str(exc),
        }), 400

@app.get('/health')
def health():
    return {"status": "ok"}


@app.route('/')
def index():
    return render_template(
        'index.html',
        coverages=list(COVERAGES),
        relations=list(RELATIONS),
        family_types=FAMILY_TYPES
    )

@app.post('/api/random-family')
def api_random_family():
    data = request.get_json(force=True)
    family_type = data.get('family_type', '1A')
    rels = random_family_from_type(
        family_type,
        data.get('gender', 'M')
    )
    return {
        'relationships': rels,
        'members': [
            mock_member(
                i + 1,
                relation,
                data.get('sum_insured', '1000000')
            )
            for i, relation in enumerate(rels)
        ]
    }

@app.post('/api/generate')
def api_generate():
    return build_payload(request.get_json(force=True))

@app.post('/download')
def download():
    payload = request.get_json(force=True)
    f = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8')
    json.dump(payload, f, indent=2, ensure_ascii=False); f.close()
    return send_file(f.name, as_attachment=True, download_name='quote_scenario.json')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))